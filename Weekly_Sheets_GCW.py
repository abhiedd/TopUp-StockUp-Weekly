import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image
from rembg import remove
import requests
import zipfile
import io
import re

st.set_page_config(layout="wide")

st.set_page_config(layout="wide")

st.markdown("""
**Required Excel Sheet Columns:**

- `Campaign Names | Asset | Grid Details | PID1 | Name1 | PID2 | Name2`

Please ensure your Excel sheet contains these columns with the exact names for proper processing.
""")
st.title("Campaigns like MagicSale, Holi, IndependenceDay(Seperate Tabs for Hubs, Excel Only, Figma S3 AmzId Columns)")

REQUIRED_COLS = ["Campaign Names", "Asset", "Grid Details", "PID1", "PID2"]

def robust_read_csv(csv_file):
    try:
        df = pd.read_csv(csv_file)
    except Exception:
        csv_file.seek(0)
        df = pd.read_csv(csv_file, encoding="latin1")
    df.columns = [str(c).strip() for c in df.columns]
    return df

def make_img_map(product_df):
    img_map = {}
    src_map = {}
    for _, row in product_df.iterrows():
        pid = str(row['MB_id']).strip()
        img_src = str(row['image_src']).strip()
        if pid and img_src and pid.lower() != 'nan' and img_src.lower() != 'nan':
            img_map[pid] = f"https://file.milkbasket.com/products/{img_src}"
            src_map[pid] = img_src
    return img_map, src_map

def make_amz_link(MB_id):
    # Replace extension with .png and build Figma S3 URL
    if not MB_id:
        return ""
    MB_id_png = re.sub(r'\.\w+$', '.png', MB_id)
    return f"https://design-figma.s3.ap-south-1.amazonaws.com/{MB_id_png}"

def fix_pid(pid):
    if pd.isna(pid) or str(pid).strip().lower() == "nan":
        return ""
    try:
        pstr = str(int(float(pid))).strip()
        return pstr
    except Exception:
        return str(pid).strip()

def clean_tab_name(campaign, asset):
    return f"{campaign.strip()} | {asset.strip()}"

def process_hub_tab(df, hub, img_map=None, src_map=None):
    df = df.copy()
    df["Campaign Names"] = df["Campaign Names"].ffill()
    df["Asset"] = df["Asset"].ffill()
    rows = []
    for _, row in df.iterrows():
        asset_val = str(row.get("Asset", "")).strip().lower()
        if asset_val in ["atc", "atc background"]:
            continue
        campaign = str(row.get("Campaign Names", "")).strip()
        asset = str(row.get("Asset", "")).strip()
        if not campaign or not asset:
            continue
        pid1 = fix_pid(row.get("PID1", ""))
        pid2 = fix_pid(row.get("PID2", ""))
        if not pid1 and not pid2:
            continue
        fg = str(row.get("Grid Details", "")).strip()
        img1 = img_map.get(pid1, "") if img_map else ""
        img2 = img_map.get(pid2, "") if img_map else ""
        src1 = src_map.get(pid1, "") if src_map else ""
        src2 = src_map.get(pid2, "") if src_map else ""
        amzid1 = make_amz_link(src1)
        amzid2 = make_amz_link(src2)
        rows.append({
            "tab": clean_tab_name(campaign, asset),
            "Hub": hub,
            "Focus Grid": fg,
            "PID1": pid1,
            "PID2": pid2,
            "Img1": img1,
            "Img2": img2,
            "AmzId1": amzid1,
            "AmzId2": amzid2
        })
    return rows

def generate_tabs_from_rows(all_rows):
    tabs = {}
    for r in all_rows:
        tabname = r["tab"]
        outrow = {k: r[k] for k in ["Hub", "Focus Grid", "PID1", "PID2", "Img1", "Img2", "AmzId1", "AmzId2"]}
        if tabname not in tabs:
            tabs[tabname] = []
        tabs[tabname].append(outrow)
    return tabs

def clean_sheet_name(name):
    return re.sub(r'[\[\]\*:/\\?]', '', str(name)).strip()[:31]

def excel_export(tabs, all_pids_tab):
    output = BytesIO()
    wb = Workbook()
    for tname, rows in tabs.items():
        ws = wb.create_sheet(title=clean_sheet_name(tname))
        df = pd.DataFrame(rows)
        df = df[["Hub", "Focus Grid", "PID1", "PID2", "Img1", "Img2", "AmzId1", "AmzId2"]]
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        bold_font = Font(bold=True)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            if r_idx == 1:
                for c_idx in range(1, len(row) + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.fill = yellow_fill
                    cell.font = bold_font
    ws2 = wb.create_sheet("All_PIDs")
    pid_df = pd.DataFrame(all_pids_tab)
    for r_idx, row in enumerate(dataframe_to_rows(pid_df, index=False, header=True), 1):
        ws2.append(row)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(output)
    output.seek(0)
    return output

def get_all_unique_pids(all_rows, img_map, src_map):
    pid_set = set()
    for r in all_rows:
        for pid_col in ["PID1", "PID2"]:
            pid = r[pid_col]
            if pid:
                pid_set.add(pid)
    all_pids = sorted(pid_set, key=lambda x: (0, int(x)) if str(x).isdigit() else (1, str(x)))
    rows = []
    for pid in all_pids:
        img_link = img_map.get(pid, "")
        src = src_map.get(pid, "")
        amzid = make_amz_link(src)
        rows.append({"PID": pid, "Img Link": img_link, "AmzID": amzid})
    return rows

def has_transparency(img_bytes):
    try:
        img = Image.open(io.BytesIO(img_bytes))
        if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
            alpha = img.getchannel("A") if "A" in img.getbands() else None
            if alpha and alpha.getextrema()[0] < 255:
                return True
        return False
    except Exception:
        return False

def remove_bg_u2net(img_bytes):
    img = Image.open(io.BytesIO(img_bytes)).convert("RGBA")
    result = remove(img)  # Uses U^2-Net by default
    img_byte_arr = io.BytesIO()
    result.save(img_byte_arr, format='PNG')
    return img_byte_arr.getvalue()

# --- UI ---

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])
product_csv = st.file_uploader("Upload product CSV (with 'MB_id' and 'image_src' columns)", type=["csv"])

img_map, src_map = {}, {}
if product_csv:
    product_df = robust_read_csv(product_csv)
    required_cols = {"MB_id", "image_src"}
    if not required_cols.issubset(product_df.columns):
        st.error("CSV must have columns 'MB_id' and 'image_src'.")
        st.stop()
    product_df["MB_id"] = product_df["MB_id"].apply(fix_pid)
    img_map, src_map = make_img_map(product_df)

all_rows = []
if uploaded_file:
    xls = pd.ExcelFile(uploaded_file)
    for tab in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=tab, header=0)
        if all(col in df.columns for col in REQUIRED_COLS):
            rows = process_hub_tab(df, tab, img_map, src_map)
            all_rows.extend(rows)

if all_rows:
    tabs = generate_tabs_from_rows(all_rows)
    all_pids_tab = get_all_unique_pids(all_rows, img_map, src_map)
    tab_names = sorted(tabs.keys())
    selected_tab = st.selectbox("Select output tab for preview/download:", tab_names)
    preview_df = pd.DataFrame(tabs[selected_tab])
    st.dataframe(preview_df)
    output = excel_export(tabs, all_pids_tab)
    st.success("✅ Excel file with multi-campaign tabs + All_PIDs ready!")
    st.download_button(
        "📥 Download Multi-Tab Excel",
        output,
        file_name="Campaign_Assetwise_MultiTab_Output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    with st.expander("Preview All_PIDs tab"):
        st.dataframe(pd.DataFrame(all_pids_tab))

    st.markdown("## Download/Process All Unique Images (All_PIDs tab)")
    if st.button("Download ALL images.zip"):
        all_img_rows = [r for r in all_pids_tab if r.get("Img Link")]
        progress = st.progress(0)
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zipf:
            for idx, item in enumerate(all_img_rows):
                pid = item["PID"]
                img_url = item["Img Link"]
                orig_filename = f"{pid}.png"
                try:
                    r = requests.get(img_url, timeout=10)
                    if r.status_code == 200:
                        img = Image.open(io.BytesIO(r.content)).convert("RGBA")
                        img_byte_arr = io.BytesIO()
                        img.save(img_byte_arr, format='PNG')
                        zipf.writestr(orig_filename, img_byte_arr.getvalue())
                except Exception:
                    continue
                progress.progress((idx + 1) / len(all_img_rows))
        zip_buffer.seek(0)
        st.success("All unique images collected! Download ready below.")
        st.download_button(
            label="Download All_PIDs_Images.zip",
            data=zip_buffer,
            file_name="All_PIDs_Images.zip",
            mime="application/zip"
        )

    if st.button("Download rembg All images.zip"):
    all_img_rows = [r for r in all_pids_tab if r.get("Img Link")]
    progress = st.progress(0)
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zipf:
        for idx, item in enumerate(all_img_rows):
            pid = item["PID"]
            img_url = item["Img Link"]
            orig_filename = f"{pid}.png"
            try:
                r = requests.get(img_url, timeout=10)
                if r.status_code == 200:
                    img_bytes = r.content
                    if has_transparency(img_bytes):
                        img_byte_arr = io.BytesIO()
                        img = Image.open(io.BytesIO(img_bytes)).convert("RGBA")
                        img.save(img_byte_arr, format='PNG')
                        zipf.writestr(orig_filename, img_byte_arr.getvalue())
                    else:
                        bg_removed_bytes = remove_bg_u2net(img_bytes)
                        zipf.writestr(orig_filename, bg_removed_bytes)
            except Exception:
                continue
            progress.progress((idx + 1) / len(all_img_rows))
    zip_buffer.seek(0)
    st.success("rembg images ready! Download below.")
    st.download_button(
        label="Download rembg_All_PIDs_Images.zip",
        data=zip_buffer,
        file_name="rembg_All_PIDs_Images.zip",
        mime="application/zip"
    )
else:
    if uploaded_file:
        st.warning("No product data found after filtering. Check if your sheet tabs/columns are correct.")
